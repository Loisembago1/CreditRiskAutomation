import os
import csv
import re
from openpyxl import load_workbook
import fitz  # PyMuPDF
from datetime import datetime

# === Paths ===
BASE_DIR = "data/ca1_cases"
CA1_ID = "CA1_20250422_0001"
CA1_PATH = os.path.join(BASE_DIR, CA1_ID)

# === Utility Functions ===

def find_case_insensitive_xlsx(dir_path, keyword):
    for fname in os.listdir(dir_path):
        name, ext = os.path.splitext(fname)
        if ext.lower() == '.xlsx' and keyword.lower() in name.lower():
            return os.path.join(dir_path, fname)
    return None

def read_dscr(path, sheet_name):
    if not path or not os.path.exists(path):
        print(f"   ⚠️ File not found at path: {path}")
        return None
    wb = load_workbook(path, data_only=True)
    print(f"   📋 Sheets in {os.path.basename(path)}: {wb.sheetnames}")
    if sheet_name not in wb.sheetnames:
        print(f"   ⚠️ Sheet '{sheet_name}' not found.")
        return None
    raw = wb[sheet_name]['C7'].value
    print(f"   🧪 Raw value at {sheet_name}!C7: {repr(raw)}")
    if raw is None:
        print("   ⚠️ Cell C7 is empty.")
        return None
    pct = raw * 100
    print(f"➡️ Extracted {sheet_name} DSCR: {pct:.2f}%")
    return pct

def decide_mcf(pct):
    if pct is None:
        return 'MCF Not Provided'
    if pct <= 18:
        return 'Premium Advantage'
    if pct <= 20:
        return 'Base Approval'
    return 'Decline'

def decide_bcf(pct):
    if pct is None:
        return 'BCF Not Provided'
    return 'Approved' if pct <= 43 else 'Decline'

def extract_fsv_from_val(pdf_path):
    print(f"\n📄 Extracting FSV details from: {os.path.basename(pdf_path)}")
    try:
        with fitz.open(pdf_path) as doc:
            text = "".join([page.get_text() for page in doc])

        text = text.replace(",", "").replace("KSH", "").replace("Ksh", "").upper()
        fsv_match = re.search(r"FORCED VALUE\s*[-:\s]*([\d]{3,})", text)
        yom_match = re.search(r"YOM\s*[:\-]*\s*([0-9]{4})", text)
        asset_type_match = re.search(r"VEHICLE TYPE\s*[:\-]*\s*([A-Z ]+)", text)

        fsv_value = int(fsv_match.group(1)) if fsv_match else None
        yom = int(yom_match.group(1)) if yom_match else None
        asset_type = asset_type_match.group(1).strip().title() if asset_type_match else None

        print(f"🚗 Asset Type: {asset_type or 'Not found'}")
        print(f"📆 Year of Manufacture: {yom or 'Not found'}")
        print(f"💰 Forced Sale Value (FSV): {fsv_value or 'Not found'}")
        return fsv_value, yom, asset_type
    except Exception as e:
        print(f"❌ Error reading {pdf_path}: {e}")
        return None, None, None
def extract_crb_info(pdf_path):
    """
    Extract Metro Score and PPI from CRB.pdf and return a credit status decision.
    """
    print(f"\n📄 Extracting CRB info from: {os.path.basename(pdf_path)}")

    try:
        with fitz.open(pdf_path) as doc:
            text = ""
            for page in doc:
                text += page.get_text()

        text = text.upper()

        # Updated patterns based on real PDF
        score_match = re.search(r"METRO-SCORE[^0-9]*([0-9]{3})", text)
        ppi_match = re.search(r"\bPPI[^A-Z0-9]*(M[1-5])\b", text)

        score = int(score_match.group(1)) if score_match else None
        ppi = ppi_match.group(1) if ppi_match else None

        print(f"💳 Metro Score: {score or 'Not found'}")
        print(f"📊 PPI Index: {ppi or 'Not found'}")

        if score is None or ppi is None:
            return score, ppi, "⚠️ Insufficient CRB Info"

        if score >= 400 and ppi in ["M1", "M2", "M3"]:
            return score, ppi, "✅ CRB Cleared"
        elif score >= 400 and ppi in ["M4", "M5"]:
            return score, ppi, "❌ CRB Risky - Manual Review"
        else:
            return score, ppi, "❌ CRB Decline"

    except Exception as e:
        print(f"❌ Error reading CRB file: {e}")
        return None, None, "❌ CRB Error"


def decide_approved_amount(fsv_value, yom, crb_verdict, olb_present):
    if not fsv_value:
        print("❌ No FSV provided. Cannot approve amount.")
        return None
    if crb_verdict != "✅ CRB Cleared":
        print("❌ CRB not cleared. Cannot approve amount.")
        return None
    if not olb_present:
        print("❌ OLB (Logbook) not attached. Cannot approve amount.")
        return None

    current_year = datetime.now().year
    age = current_year - yom if yom else None

    if yom and age <= 8:
        percentage = 0.70
    elif yom and age <= 12:
        percentage = 0.60
    else:
        print("❌ Vehicle too old for approval.")
        return None

    approved_amount = round(fsv_value * percentage, -3)
    print(f"✅ Approved Amount: KES {approved_amount:,} ({percentage*100:.0f}% of FSV)")
    return approved_amount

# === Main Program ===

def main():
    if not os.path.exists(CA1_PATH):
        print("❌ CA1 folder not found.")
        return

    print(f"\n📂 Scanning files in: {CA1_PATH}\n")
    for f in os.listdir(CA1_PATH):
        print(f"✅ Found: {f}")

    mcf_path = find_case_insensitive_xlsx(CA1_PATH, 'mcf')
    bcf_path = find_case_insensitive_xlsx(CA1_PATH, 'bcf')
    val_path = os.path.join(CA1_PATH, "VAL.pdf")
    crb_path = os.path.join(CA1_PATH, "CRB.pdf")

    mcf_pct = read_dscr(mcf_path, 'Output Template')
    bcf_pct = read_dscr(bcf_path, 'Output Template')
    mcf_decision = decide_mcf(mcf_pct)
    bcf_decision = decide_bcf(bcf_pct)
    final_verdict = f"{mcf_decision} + {bcf_decision}"

    print(f"\n🎯 MCF Decision: {mcf_decision}")
    print(f"🎯 BCF Decision: {bcf_decision}")
    print(f"\n🧾 Final Verdict: {final_verdict}")

    fsv_value, asset_year, asset_type = extract_fsv_from_val(val_path)
    crb_score, crb_ppi, crb_verdict = extract_crb_info(crb_path)

    olb_present = any("olb" in f.lower() for f in os.listdir(CA1_PATH))
    approved_amount = decide_approved_amount(fsv_value, asset_year, crb_verdict, olb_present)

    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, "results.csv")
    write_headers = not os.path.exists(output_file)

    headers = [
        "Case ID", "MCF DSCR (%)", "MCF Decision",
        "BCF DSCR (%)", "BCF Decision", "Final Verdict",
        "FSV Value", "Asset Year", "Asset Type",
        "CRB Score", "PPI", "CRB Verdict", "Approved Amount"
    ]

    row = [
        CA1_ID, mcf_pct, mcf_decision,
        bcf_pct, bcf_decision, final_verdict,
        fsv_value, asset_year, asset_type,
        crb_score, crb_ppi, crb_verdict,
        approved_amount
    ]

    with open(output_file, mode="a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if write_headers:
            writer.writerow(headers)
        writer.writerow(row)

    print(f"💳 CRB Decision: {crb_verdict}")
    print(f"\n💾 Results saved to: {output_file}")

# === Run ===
if __name__ == "__main__":
    main()
